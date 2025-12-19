from django.shortcuts import render, redirect
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.contrib import messages
from .models import Producto, Proveedor, Movimiento
from .forms import ProductoForm, ProveedorForm, MovimientoForm
import datetime
import csv
import openpyxl
import openpyxl
from openpyxl.styles import Font
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json 

def almacen_dashboard(request):
    productos = Producto.objects.all()
    proveedores = Proveedor.objects.all()
    movimientos = Movimiento.objects.order_by('-fecha')

    # ---------------------------
    # FILTROS
    # ---------------------------
    producto_id = request.GET.get('producto')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_mov = request.GET.get('tipo_mov')  # ingreso/salida

    filtro = Q()
    if producto_id:
        filtro &= Q(producto_id=producto_id)
    if fecha_inicio and fecha_fin:
        filtro &= Q(fecha__range=[fecha_inicio, fecha_fin])
    if tipo_mov in ['ingreso', 'salida']:
        filtro &= Q(tipo__startswith=tipo_mov)

    movimientos = movimientos.filter(filtro)

    # ---------------------------
    # FORMULARIOS
    # ---------------------------
    producto_form = ProductoForm()
    proveedor_form = ProveedorForm()
    movimiento_form = MovimientoForm()

    if request.method == "POST":
        # Crear producto individual
        if 'producto_submit' in request.POST:
            producto_form = ProductoForm(request.POST)
            if producto_form.is_valid():
                producto_form.save()
                messages.success(request, "Producto creado correctamente.")
                return redirect('almacen_dashboard')


        elif 'carga_masiva_productos' in request.POST:
            archivo = request.FILES.get('archivo_excel')

            if not archivo:
                messages.error(request, "Debe seleccionar un archivo Excel.")
                return redirect('almacen_dashboard')

            import openpyxl
            from decimal import Decimal, InvalidOperation

            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
            except Exception as e:
                messages.error(request, f"No se pudo leer el archivo: {str(e)}")
                return redirect('almacen_dashboard')

            # Normalizar encabezados
            headers = [str(c.value).strip().lower().replace(" ", "_") for c in ws[1]]

            expected = ['nombre', 'referencia', 'utilidad', 'precio_venta', 'ean']

            if headers != expected:
                messages.error(
                    request,
                    "Encabezados inválidos.\n"
                    f"Esperados: {expected}\n"
                    f"Recibidos: {headers}"
                )
                return redirect('almacen_dashboard')


            creados = 0
            errores = []

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not any(row):
                        continue

                    nombre, referencia, utilidad_raw, precio_raw, ean_raw = row

                    if not nombre or not referencia:
                        raise ValueError("Nombre o referencia vacíos")

                    # UTILIDAD → texto tal cual (solo limpieza básica)
                    utilidad = str(utilidad_raw).strip() if utilidad_raw else None

                    # PRECIO → sí o sí Decimal
                    precio_venta = Decimal(str(precio_raw).strip().replace(",", "."))

                    # EAN → evitar notación científica
                    ean = str(ean_raw).split(".")[0] if ean_raw else None

                    Producto.objects.create(
                        nombre=str(nombre).strip(),
                        referencia=str(referencia).strip(),
                        utilidad=utilidad,
                        precio_venta=precio_venta,
                        ean=ean
                    )

                    creados += 1

                except Exception as e:
                    errores.append(f"Fila {i}: {e}")



            if errores:
                filas = [e.split(":")[0] for e in errores]

                messages.warning(
                    request,
                    f"Se crearon {creados} productos.<br>"
                    f"Errores en: {', '.join(filas)}"
                )

                for e in errores:
                    print("⚠️", e)
            else:
                messages.success(
                    request,
                    f"Se cargaron {creados} productos correctamente."
                )

            return redirect('almacen_dashboard')



        # Crear proveedor
        elif 'proveedor_submit' in request.POST:
            proveedor_form = ProveedorForm(request.POST)
            if proveedor_form.is_valid():
                proveedor_form.save()
                messages.success(request, "Proveedor creado correctamente.")
                return redirect('almacen_dashboard')

        # Crear movimiento
        elif 'movimiento_submit' in request.POST:
            movimiento_form = MovimientoForm(request.POST)
            if movimiento_form.is_valid():
                movimiento_form.save()
                messages.success(request, "Movimiento registrado correctamente.")
                return redirect('almacen_dashboard')

        # Exportar CSV
        elif 'export_csv' in request.POST:
            return export_movimientos_csv(movimientos)

        
        # Carga masiva de facturas
        elif 'carga_masiva_factura' in request.POST:
            archivo = request.FILES.get('archivo_factura')
            if not archivo:
                messages.error(request, "Debe seleccionar un archivo.")
                return redirect('almacen_dashboard')

            import openpyxl
            from openpyxl.styles import Font
            from django.http import HttpResponse
            from decimal import Decimal, InvalidOperation

            try:
                wb = openpyxl.load_workbook(archivo)
                ws = wb.active
            except Exception as e:
                messages.error(request, f"No se pudo leer el archivo: {str(e)}")
                return redirect('almacen_dashboard')

            expected_headers = ['referencia', 'cantidad', 'precio_unitario', 'proveedor', 'factura_referencia']
            actual_headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]

            print("👉 Encabezados detectados:", actual_headers)  # Debug opcional

            if actual_headers != expected_headers:
                messages.error(request, f"Encabezados inválidos. Se esperaban: {expected_headers}")
                return redirect('almacen_dashboard')

            movimientos_creados = 0
            productos_no_encontrados = []
            proveedores_no_encontrados = []
            errores = []

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    referencia_o_nombre, cantidad_raw, precio_unitario_raw, proveedor_nombre, factura_ref = row

                    if not referencia_o_nombre:
                        raise ValueError("Referencia o nombre vacío en la fila")

                    # Buscar producto por referencia exacta o por nombre similar
                    ref_str = str(referencia_o_nombre).strip()
                    producto = Producto.objects.filter(referencia__iexact=ref_str).first()
                    if not producto:
                        producto = Producto.objects.filter(nombre__icontains=ref_str).first()

                    if not producto:
                        productos_no_encontrados.append({
                            'fila': i,
                            'referencia_o_nombre': ref_str,
                            'factura': factura_ref or ''
                        })
                        continue

                    # Buscar proveedor (NO crear si no existe)
                    proveedor = None
                    if proveedor_nombre:
                        proveedor = Proveedor.objects.filter(nombre__iexact=str(proveedor_nombre).strip()).first()
                        if not proveedor:
                            proveedores_no_encontrados.append({
                                'fila': i,
                                'proveedor': proveedor_nombre,
                                'referencia_o_nombre': ref_str
                            })
                            continue  # no crea movimiento si proveedor no existe

                    # Validar cantidad y precio
                    try:
                        cantidad = int(cantidad_raw)
                    except (TypeError, ValueError):
                        raise ValueError("Cantidad inválida (debe ser un número entero)")

                    try:
                        precio_unitario = Decimal(str(precio_unitario_raw))
                    except (TypeError, InvalidOperation):
                        raise ValueError("Precio unitario inválido (debe ser numérico)")

                    # Crear movimiento
                    Movimiento.objects.create(
                        producto=producto,
                        tipo='ingreso_compra',
                        cantidad=cantidad,
                        proveedor=proveedor,
                        precio_unitario=precio_unitario,
                        factura_referencia=str(factura_ref).strip() if factura_ref else None
                    )

                    movimientos_creados += 1

                except Exception as e:
                    errores.append({'fila': i, 'error': str(e)})

            # Debug opcional
            print(f"✅ Movimientos creados: {movimientos_creados}")
            print(f"❌ Productos no encontrados: {productos_no_encontrados}")
            print(f"⚠️ Proveedores no encontrados: {proveedores_no_encontrados}")
            print(f"⚠️ Errores detectados: {errores}")

            # Generar log si hay errores
            if errores or productos_no_encontrados or proveedores_no_encontrados:
                log_wb = openpyxl.Workbook()
                log_ws = log_wb.active
                log_ws.title = "Log Carga Factura"
                log_ws.append(["Tipo", "Fila", "Referencia/Nombre", "Proveedor", "Factura", "Detalle"])

                for prod in productos_no_encontrados:
                    log_ws.append([
                        "Producto no encontrado",
                        prod["fila"],
                        prod["referencia_o_nombre"],
                        "-",
                        prod.get("factura", "-"),
                        "-"
                    ])
                for prov in proveedores_no_encontrados:
                    log_ws.append([
                        "Proveedor no encontrado",
                        prov["fila"],
                        prov.get("referencia_o_nombre", "-"),
                        prov["proveedor"],
                        "-",
                        "-"
                    ])
                for err in errores:
                    log_ws.append([
                        "Error",
                        err["fila"],
                        "-",
                        "-",
                        "-",
                        err["error"]
                    ])

                for cell in log_ws[1]:
                    cell.font = Font(bold=True)

                response = HttpResponse(
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response["Content-Disposition"] = 'attachment; filename="log_carga_factura.xlsx"'
                log_wb.save(response)
                return response

            # Mensajes finales
            if movimientos_creados:
                messages.success(request, f"Se crearon {movimientos_creados} movimientos correctamente.")
            else:
                messages.info(request, "No se crearon movimientos (productos o proveedores no encontrados, o archivo vacío).")

            return redirect('almacen_dashboard')



    # ---------------------------
    # AGRUPAR MOVIMIENTOS POR PRODUCTO
    # ---------------------------
    movimientos_por_producto = {}
    for producto in productos:
        movimientos_producto = movimientos.filter(producto=producto)
        ingresos = movimientos_producto.filter(tipo__startswith='ingreso')
        salidas = movimientos_producto.filter(tipo__startswith='salida')

        stock_ingresos = ingresos.aggregate(total=Sum('cantidad'))['total'] or 0
        stock_salidas = salidas.aggregate(total=Sum('cantidad'))['total'] or 0
        stock_actual = stock_ingresos - stock_salidas

        valor_ingresos = sum([m.valor_total() for m in ingresos])
        valor_salidas = sum([m.valor_total() for m in salidas])
        valor_stock = valor_ingresos - valor_salidas

        movimientos_por_producto[producto] = {
            "ingresos": ingresos,
            "salidas": salidas,
            "stock_actual": stock_actual,
            "valor_stock": valor_stock
        }

    context = {
        "productos": productos,
        "proveedores": proveedores,
        "movimientos_por_producto": movimientos_por_producto,
        "producto_form": producto_form,
        "proveedor_form": proveedor_form,
        "movimiento_form": movimiento_form,
        "fecha_inicio": fecha_inicio or "",
        "fecha_fin": fecha_fin or "",
        "producto_id": producto_id or "",
        "tipo_mov": tipo_mov or "",
    }

    return render(request, "almacen/almacen_tree_stock.html", context)


# ------------------------------------------------------
# FUNCIÓN AUXILIAR: Exportar CSV
# ------------------------------------------------------
def export_movimientos_csv(movimientos):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="movimientos_{datetime.date.today()}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Producto', 'Tipo', 'Cantidad', 'Proveedor', 'Fecha', 'Precio Unitario', 'Valor Total'])
    for m in movimientos:
        writer.writerow([
            m.producto.nombre,
            m.get_tipo_display(),
            m.cantidad,
            m.proveedor.nombre if m.proveedor else "-",
            m.fecha,
            m.precio_unitario,
            m.valor_total()
        ])
    return response

def editar_producto(request, id):
    """Actualiza los datos de un producto en tiempo real (AJAX)."""
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            producto = Producto.objects.get(pk=id)

            producto.nombre = data.get("nombre", producto.nombre)
            producto.referencia = data.get("referencia", producto.referencia)
            producto.utilidad = data.get("utilidad", producto.utilidad)
            producto.precio_venta = data.get("precio_venta", producto.precio_venta)
            producto.ean = data.get("ean", producto.ean)
            producto.save()

            return JsonResponse({"success": True})
        except Producto.DoesNotExist:
            return JsonResponse({"success": False, "error": "Producto no encontrado."}, status=404)
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    return JsonResponse({"success": False, "error": "Método no permitido."}, status=405)
